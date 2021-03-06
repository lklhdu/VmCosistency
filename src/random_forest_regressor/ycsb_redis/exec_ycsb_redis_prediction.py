from sklearn.ensemble import RandomForestRegressor
import pymysql
from pymysql.cursors import DictCursor
import numpy as np
import openpyxl
import math
from openpyxl.styles import PatternFill
from sklearn import metrics
from sklearn.model_selection import GridSearchCV
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import Ridge
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import PolynomialFeatures
from sklearn import metrics
from sklearn.feature_selection import SelectKBest, f_regression, mutual_info_regression
import warnings
from sklearn.tree import DecisionTreeRegressor

warnings.filterwarnings('ignore')


#  MAPE和SMAPE
def mape(y_true, y_pred):
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_pred.astype(np.float64)
    y_true.astype(np.float64)
    return np.mean(np.abs((y_pred - y_true) / y_true)) * 100


def smape(y_true, y_pred):
    y_true = np.array(y_true)
    y_pred = np.array(y_pred)
    y_pred.astype(np.float)
    y_true.astype(np.float)
    return 2.0 * np.mean(np.abs(y_pred - y_true) / (np.abs(y_pred) + np.abs(y_true))) * 100


def ycsb_redis_prediction():
    train_data = []
    train_data_target = []
    test_data = []
    test_data_target = []
    attributes = ["server_type", "cpu_number", "reciprecal_real", "memory_count", "triad"]
    print(attributes)

    # # 列表中元素的数据类型为字典
    # list_regressor_data = []
    # db = pymysql.connect("localhost", "root", "me521..", "vmconsistency")
    # # 查询返回字典值
    # cursor = db.cursor(DictCursor)
    # cursor.execute('select * from ycsb_run_redis')
    #
    # # 查询后的字段名称可以有cursor.description
    # # for col in cursor.description:
    # #     print(col)
    #
    # ycsb_redis_recordings = cursor.fetchall()
    #
    # for ycsb_redis_recording in ycsb_redis_recordings:
    #     regressor_data = {}
    #     # 内存
    #     stream_select_sql = "select * from stream where cpu_number=" + str(
    #         ycsb_redis_recording['cpu_number']) + " and cpu_frequency=" + str(
    #         ycsb_redis_recording['cpu_frequency']) + " and memory_count=" + str(
    #         ycsb_redis_recording['memory_count']) + " and server_type=" + str(ycsb_redis_recording['server_type'])
    #     cursor.execute(stream_select_sql)
    #     stream_select_result = cursor.fetchall()
    #     # 磁盘I/O读
    #     fio_read_select_sql = "select * from fio_read where cpu_number=" + str(
    #         ycsb_redis_recording['cpu_number']) + " and cpu_frequency=" + str(
    #         ycsb_redis_recording['cpu_frequency']) + " and memory_count=" + str(
    #         ycsb_redis_recording['memory_count']) + " and server_type=" + str(ycsb_redis_recording['server_type'])
    #     cursor.execute(fio_read_select_sql)
    #     fio_read_select_result = cursor.fetchall()
    #     # 磁盘I/O写
    #     fio_write_select_sql = "select * from fio_write where cpu_number=" + str(
    #         ycsb_redis_recording['cpu_number']) + " and cpu_frequency=" + str(
    #         ycsb_redis_recording['cpu_frequency']) + " and memory_count=" + str(
    #         ycsb_redis_recording['memory_count']) + " and server_type=" + str(ycsb_redis_recording['server_type'])
    #     cursor.execute(fio_write_select_sql)
    #     fio_write_select_result = cursor.fetchall()
    #
    #     linpack_select_sql = "select * from linpack where cpu_number=" + str(
    #         ycsb_redis_recording['cpu_number']) + " and cpu_frequency=" + str(
    #         ycsb_redis_recording['cpu_frequency']) + " and memory_count=" + str(
    #         ycsb_redis_recording['memory_count']) + " and server_type=" + str(ycsb_redis_recording['server_type'])
    #     cursor.execute(linpack_select_sql)
    #     linpack_select_result = cursor.fetchall()
    #
    #     # CPU计算
    #     pi5000_select_sql = "select * from pi5000 where cpu_number=" + str(
    #         ycsb_redis_recording['cpu_number']) + " and cpu_frequency=" + str(
    #         ycsb_redis_recording['cpu_frequency']) + " and memory_count=" + str(
    #         ycsb_redis_recording['memory_count']) + " and server_type=" + str(ycsb_redis_recording['server_type'])
    #     cursor.execute(pi5000_select_sql)
    #     pi5000_select_result = cursor.fetchall()
    #
    #     # kernel_select_sql = "select * from kernel where cpu_number=" + str(
    #     #     ycsb_redis_recording['cpu_number']) + " and cpu_frequency=" + str(
    #     #     ycsb_redis_recording['cpu_frequency']) + " and memory_count=" + str(
    #     #     ycsb_redis_recording['memory_count']) + " and server_type=" + str(ycsb_redis_recording['server_type'])
    #     # cursor.execute(kernel_select_sql)
    #     # kernel_select_result = cursor.fetchall()
    #
    #     if (len(stream_select_result) == 1 and len(linpack_select_result) == 1 and len(fio_read_select_result) == 1 and len(
    #             fio_write_select_result) == 1 and len(pi5000_select_result) == 1):
    #         regressor_data.update(stream_select_result[0])
    #         regressor_data.update(fio_read_select_result[0])
    #         regressor_data.update(fio_write_select_result[0])
    #         regressor_data.update(pi5000_select_result[0])
    #         regressor_data.update(linpack_select_result[0])
    #         # regressor_data.update(kernel_select_result[0])
    #         regressor_data.update(ycsb_redis_recording)
    #         list_regressor_data.append(regressor_data)
    #         # print(regressor_data)
    #
    #
    #
    # np_list_regressor_data = np.array(list_regressor_data)
    # np.save("np_list_regressor_data.npy", np_list_regressor_data)

    np_list_regressor_data = np.load("E:\\PyCharm-workspace\\VmCosistency\src\\random_forest_regressor\\ycsb_redis\\np_list_regressor_data.npy")

    print(len(np_list_regressor_data))

    labels = []
    for regressor_data in np_list_regressor_data:
        labels.append(float(regressor_data['throughput']))

    train, test, train_data_target, test_data_target = train_test_split(np_list_regressor_data, labels, test_size=0.95)

    for iter_train in train:
        data = []
        for attribute in attributes:
            data.append(iter_train[attribute])
        train_data.append(data)

    for iter_test in test:
        data = []
        for attribute in attributes:
            data.append(iter_test[attribute])
        # print(data)
        test_data.append(data)

    np_train_data = np.array(train_data)
    np_test_data = np.array(test_data)
    np_test_data_target = np.array(test_data_target)
    np_train_data_target = np.array(train_data_target)

    np_train_data.astype(np.float)
    np_test_data.astype(np.float)
    np_test_data_target.astype(np.float)
    np_train_data_target.astype(np.float)

    poly_feature = PolynomialFeatures(degree=1)
    poly_train_data = poly_feature.fit_transform(np_train_data)
    poly_test_data = poly_feature.fit_transform(np_test_data)

    r2_list = []
    MAPE_list = []

    # 线性回归
    lin_reg = LinearRegression()
    lin_reg.fit(poly_train_data, train_data_target)
    predict_result = lin_reg.predict(poly_test_data)
    # print(lin_reg.coef_)
    r2 = metrics.r2_score(test_data_target, predict_result)
    MAPE = mape(test_data_target, predict_result)
    r2_list.append(r2)
    MAPE_list.append(MAPE)


    # 随机森林回归
    rfr = RandomForestRegressor()
    rfr.fit(poly_train_data, train_data_target)
    predict_result = rfr.predict(poly_test_data)
    r2 = metrics.r2_score(test_data_target, predict_result)
    MAPE = mape(test_data_target, predict_result)
    r2_list.append(r2)
    MAPE_list.append(MAPE)

    # 决策树回归
    dtr = DecisionTreeRegressor()
    dtr.fit(poly_train_data, train_data_target)
    predict_result = dtr.predict(poly_test_data)
    r2 = metrics.r2_score(test_data_target, predict_result)
    MAPE = mape(test_data_target, predict_result)
    r2_list.append(r2)
    MAPE_list.append(MAPE)

    # print("MSE 均方误差:   " + str(metrics.mean_squared_error(test_data_target, predict_result)))
    # print("MAE  平均绝对误差:  " + str(metrics.mean_absolute_error(test_data_target, predict_result)))
    print("R2 决定系数:   " + str(metrics.r2_score(test_data_target, predict_result)))
    print("RMSE 均方根误差  " + str(np.sqrt(metrics.mean_squared_error(test_data_target, predict_result))))
    # print("MAPE 平均绝对百分比误差  " + str(mape(test_data_target, predict_result)))
    # print("SMAPE 对称平均绝对百分比误差  " + str(smape(test_data_target, predict_result)))

    row = 1
    col = 1
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for column_name in attributes:
        sheet.cell(row, col, column_name)
        col += 1
    sheet.cell(row, col, "真值")
    col += 1
    sheet.cell(row, col, "预测值")
    col += 1
    sheet.cell(row, col, "预测误差百分比")
    col += 1
    sheet.cell(row, col, "预测误差")

    unqualified_number = 0
    for index in range(0, len(test_data)):
        row += 1
        for col in range(0, len(test_data[index])):
            sheet.cell(row, col + 1, test_data[index][col])
        col = len(test_data[index]) + 1
        sheet.cell(row, col, test_data_target[index])
        col += 1
        sheet.cell(row, col, predict_result[index])
        error = predict_result[index] - float(test_data_target[index])
        errorPercent = error / float(test_data_target[index]) * 100
        col += 1
        fill = PatternFill("solid", fgColor="1874CD")
        sheet.cell(row, col, errorPercent)
        col += 1
        sheet.cell(row, col, error)

        if math.fabs(errorPercent) > 10:
            sheet.cell(row, col).fill = fill
            sheet.cell(row, col - 1).fill = fill
            sheet.cell(row, col - 2).fill = fill
            unqualified_number += 1

    print(float(unqualified_number) / len(test_data_target))
    workbook.save("ycsb_redis_data.xlsx")
    return r2_list, MAPE_list


if __name__ == '__main__':
    r2, MAPE = ycsb_redis_prediction()
    print()
    print()
    print(r2)
    print(MAPE)
